# -*- coding: utf-8 -*-
"""
심평원 공개자료(엑셀) → 폴백 병원 명부 NDJSON.

★ 왜 필요한가.
  홈페이지 첫 화면의 병원 조회가 행정안전부 API 하나에 걸려 있었고, 2026-07-27 그
  API가 "성공했고 결과는 0건"을 15분간 뿜는 동안 첫 화면 전체가 죽었다.
  행안부가 죽어도 병원을 찾아낼 **우리 소유의 명부**가 필요하다.

입력 (건강보험심사평가원 공개 파일, 이미 보유):
  · hospital_info.xlsx  — 병원정보 (암호화요양기호·요양기관명·주소·전화·개설일자)
  · subjects.xlsx       — 의료기관별 진료과목 (암호화요양기호로 조인)

출력:
  NDJSON 1행 = clinic_directory 1행. 업로드는 import-clinic-directory.mjs 가 맡는다.

식별자:
  'hira:' + sha256(암호화요양기호)[:16]
  · 원본 요양기호는 80자라 API 입력 상한(60자)을 넘고, 그대로 노출할 이유도 없다.
  · 해시는 결정적이라 자료를 다시 올려도 같은 병원은 같은 식별자를 유지한다
    (재적재로 기존 리드·리포트의 mng_no 가 고아가 되지 않는다).

사용:
  python scripts/export-clinic-directory.py \
      --hospitals "C:/Users/PC/vox-pilot/hira/hospital_info.xlsx" \
      --subjects  "C:/Users/PC/vox-pilot/hira/subjects.xlsx" \
      --out       "scripts/.cache/clinic-directory.ndjson" \
      --version   2026Q1
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook

# 시·도명 → 지역번호. 심평원 원자료의 전화번호는 지역번호가 빠진 행이 섞여 있다.
AREA_CODES = {
    "서울": "02", "부산": "051", "대구": "053", "인천": "032", "광주": "062",
    "대전": "042", "울산": "052", "세종": "044", "경기": "031", "강원": "033",
    "충북": "043", "충남": "041", "전북": "063", "전남": "061", "경북": "054",
    "경남": "055", "제주": "064",
}

# 시·도 축약형 → 주소 표기(정식 명칭). 지역 필터가 주소 부분일치라 표기를 맞춰야 한다.
PROVINCE_FULL = {
    "서울": "서울특별시", "부산": "부산광역시", "대구": "대구광역시", "인천": "인천광역시",
    "광주": "광주광역시", "대전": "대전광역시", "울산": "울산광역시", "세종": "세종특별자치시",
    "경기": "경기도", "강원": "강원특별자치도", "충북": "충청북도", "충남": "충청남도",
    "전북": "전북특별자치도", "전남": "전라남도", "경북": "경상북도", "경남": "경상남도",
    "제주": "제주특별자치도",
}

# 우리가 쓰는 진료과목 어휘로 맞춘다 (registry.ts 의 SPECIALTY_ALIASES 와 동일 정책).
SPECIALTY_ALIASES = {
    "소아과": "소아청소년과",
    "비뇨기과": "비뇨의학과",
    "정신과": "정신건강의학과",
    "신경정신과": "정신건강의학과",
}


def norm_name(value: str) -> str:
    """공백 제거 + 소문자 — name_norm 의 정본 (registry.ts normalizeClinicName 과 동일)."""
    return re.sub(r"\s+", "", value or "").lower()


def directory_id(ykiho: str) -> str:
    digest = hashlib.sha256(ykiho.encode("utf-8")).hexdigest()[:16]
    return f"hira:{digest}"


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def split_address(address: str, province_hint: str) -> tuple[str, str]:
    """주소 → (시·도, 구·군). 실패하면 빈 문자열."""
    parts = [p for p in (address or "").split() if p]
    province = parts[0] if parts else PROVINCE_FULL.get(province_hint, province_hint)
    district = ""
    for p in parts[1:]:
        if p.endswith(("구", "군")):
            district = p
            break
    if not district:
        for p in parts[1:]:
            if p.endswith("시"):
                district = p
                break
    return province, district


def normalize_phone(phone: str, province_hint: str) -> str:
    """지역번호가 빠진 번호에 시·도 기준 지역번호를 붙인다. 판단이 안 서면 원문 유지."""
    value = (phone or "").strip()
    if not value:
        return ""
    if value.startswith("0"):
        return value
    code = AREA_CODES.get((province_hint or "")[:2])
    if not code:
        return value
    return f"{code}-{value}"


def load_subjects(path: str) -> dict[str, list[str]]:
    """암호화요양기호 → 진료과목 목록(전문의 수 내림차순)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)
    bucket: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for row in rows:
        if not row or not row[0]:
            continue
        ykiho = cell_text(row[0])
        subject = cell_text(row[3]) if len(row) > 3 else ""
        if not ykiho or not subject:
            continue
        try:
            specialists = int(row[4]) if len(row) > 4 and row[4] is not None else 0
        except (TypeError, ValueError):
            specialists = 0
        bucket[ykiho].append((specialists, SPECIALTY_ALIASES.get(subject, subject)))
    wb.close()

    out: dict[str, list[str]] = {}
    for ykiho, items in bucket.items():
        # 전문의 수가 많은 과목이 그 병원의 실제 간판이다 (안과 3명 vs 내과 0명 → 안과).
        items.sort(key=lambda x: -x[0])
        seen: set[str] = set()
        ordered: list[str] = []
        for _, subject in items:
            if subject in seen:
                continue
            seen.add(subject)
            ordered.append(subject)
        out[ykiho] = ordered
    return out


def pick_specialty(subjects: list[str], institution_type: str) -> str:
    if institution_type == "치과의원":
        return "치과"
    if institution_type in ("한의원", "한방병원"):
        return "한의원"
    return subjects[0] if subjects else ""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--hospitals", required=True)
    parser.add_argument("--subjects", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--version", default="")
    parser.add_argument(
        "--province",
        default="",
        help="시·도명으로 제한 (예: 대구). 비우면 전국.",
    )
    args = parser.parse_args()

    subjects_by_ykiho = load_subjects(args.subjects)

    wb = load_workbook(args.hospitals, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)

    written = 0
    skipped = 0
    seen_ids: set[str] = set()

    with open(args.out, "w", encoding="utf-8") as fp:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                skipped += 1
                continue
            ykiho = cell_text(row[0])
            name = cell_text(row[1])
            institution_type = cell_text(row[3])
            province_hint = cell_text(row[5])
            address = cell_text(row[10])
            phone = cell_text(row[11])
            opened = cell_text(row[13])

            if not ykiho or not name:
                skipped += 1
                continue
            if args.province and province_hint[:2] != args.province[:2]:
                continue

            mng_no = directory_id(ykiho)
            # 같은 요양기호가 두 번 나오면 뒤엣것은 버린다(PK 충돌 방지).
            if mng_no in seen_ids:
                skipped += 1
                continue
            seen_ids.add(mng_no)

            province, region = split_address(address, province_hint)
            subject_list = subjects_by_ykiho.get(ykiho, [])

            fp.write(
                json.dumps(
                    {
                        "mng_no": mng_no,
                        "name": name,
                        "name_norm": norm_name(name),
                        "road_address": address,
                        "province": province,
                        "region": region,
                        "institution_type": institution_type,
                        "specialty": pick_specialty(subject_list, institution_type),
                        "subjects": subject_list,
                        "phone": normalize_phone(phone, province_hint),
                        "opened_on": opened[:10] if opened else None,
                        "source": "hira",
                        "source_version": args.version,
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )
            written += 1

    wb.close()
    print(f"written={written} skipped={skipped} out={args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
